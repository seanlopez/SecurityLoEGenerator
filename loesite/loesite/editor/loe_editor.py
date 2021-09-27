from openpyxl import load_workbook


class loe_editor(object):
    def __init__(self, form_dict, workbook_file, sheet_name, buffer_cell_line, value_range):
        self.portal_form = form_dict
        self.wb = load_workbook(workbook_file)
        self.ws = self.wb[sheet_name]
        self.ws["C3"] = form_dict["customer_name"]   # fill the customer name in sheet
        self.buffer_cell_line = buffer_cell_line
        self.value_range = value_range

# Stealtwatch Editor
    
    def stw_requirement_phase_editor(self):
        '''
                according to the portal form value to fill in the workdays in LOE spreadsheet
                :return:
        '''
        workshop_days = 0.5
        document_creation_days = 3

        if "endpoint" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
        if "slic" in self.portal_form.keys():
            workshop_days = workshop_days + 0
        if "sal" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
        if "3rd" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
        if "api" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5

        # fill the workshop days

        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"])/100
            self.ws["D17"] = workshop_days * leader_effort
            self.ws["E17"] = workshop_days * (1 - leader_effort)

            self.ws["D18"] = document_creation_days * leader_effort
            self.ws["E18"] = document_creation_days * (1 - leader_effort)
        else:
            self.ws["E17"] = workshop_days
            self.ws["E18"] = document_creation_days

    def stw_design_phase_editor(self):
        '''
                according to the portal form value to fill in the workdays in LOE spreadsheet
                :return:
        '''
        workshop_days = 0.5
        document_creation_days = 4

        if "endpoint" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
            document_creation_days = document_creation_days + 0.5
        if "slic" in self.portal_form.keys():
            workshop_days = workshop_days + 0
            document_creation_days = document_creation_days + 0
        if "sal" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
            document_creation_days = document_creation_days + 0.5
        if "3rd" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
            document_creation_days = document_creation_days + 0.5
        if "api" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
            document_creation_days = document_creation_days + 0.5

        # fill the value to the spreadsheet
        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D21"] = workshop_days * leader_effort
            self.ws["E21"] = workshop_days * (1 - leader_effort)

            self.ws["D22"] = document_creation_days * leader_effort
            self.ws["E22"] = document_creation_days * (1 - leader_effort)
        else:
            self.ws["E21"] = workshop_days
            self.ws["E22"] = document_creation_days

    def stw_nip_phase_editor(self):
        '''
                according to the portal form value to fill in the workdays in LOE spreadsheet
                :return:
        '''
        workshop_days = 0.5
        document_creation_days = 4

        if "endpoint" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
            document_creation_days = document_creation_days + 0.5
        if "slic" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
            document_creation_days = document_creation_days + 0.5
        if "sal" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
            document_creation_days = document_creation_days + 0.5
        if "3rd" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
            document_creation_days = document_creation_days + 0.5
        if "api" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
            document_creation_days = document_creation_days + 0.5
        if "udpredirector" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
            document_creation_days = document_creation_days + 0.5
        if "fs" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
            document_creation_days = document_creation_days + 0.5
        if "iseintegration" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
            document_creation_days = document_creation_days + 0.5

        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D28"] = workshop_days * leader_effort
            self.ws["E28"] = workshop_days * (1 - leader_effort)

            self.ws["D29"] = document_creation_days * leader_effort
            self.ws["E29"] = document_creation_days * (1 - leader_effort)
        else:
            self.ws["E28"] = workshop_days
            self.ws["E29"] = document_creation_days

    def stw_nrfu_phase_editor(self):
        '''
            according to the portal form value to fill in the workdays in LOE spreadsheet
            :return:
        '''
        document_creation_days = 3

        if "endpoint" in self.portal_form.keys():
            document_creation_days = document_creation_days + 0.5
        if "slic" in self.portal_form.keys():
            document_creation_days = document_creation_days + 0.5
        if "sal" in self.portal_form.keys():
            document_creation_days = document_creation_days + 0.5
        if "3rd" in self.portal_form.keys():
            document_creation_days = document_creation_days + 0.5
        if "api" in self.portal_form.keys():
            document_creation_days = document_creation_days + 0.5
        if "iseintegration" in self.portal_form.keys():
            document_creation_days = document_creation_days + 0.5

        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D35"] = document_creation_days * leader_effort
            self.ws["E35"] = document_creation_days * (1 - leader_effort)
        else:
            self.ws["E35"] = document_creation_days

    def stw_lab_testing_phase_editor(self):
        '''
                        according to the portal form value to fill in the workdays in LOE spreadsheet
                        :return:
        '''
        lab_building_days = 0.5
        lab_test_days = 0.5

        leader_effort = 0
        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100

        if "endpoint" in self.portal_form.keys():
            lab_building_days = lab_building_days + 0.5
            lab_test_days = lab_test_days + 0.5
        if "slic" in self.portal_form.keys():
            lab_building_days = lab_building_days + 0.5
            lab_test_days = lab_test_days + 0.5
        if "sal" in self.portal_form.keys():
            lab_building_days = lab_building_days + 1
            lab_test_days = lab_test_days + 1
        if "3rd" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D44"] = 5 * leader_effort
                self.ws["E44"] = 5 * (1 - leader_effort)
            else:
                self.ws["E44"] = 5
        if "api" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D45"] = 5 * leader_effort
                self.ws["E45"] = 5 * (1 - leader_effort)
            else:
                self.ws["E45"] = 5

        if "udpredirector" in self.portal_form.keys():
            lab_building_days = lab_building_days + 0.5
            lab_test_days = lab_test_days + 0.5
        if "fs" in self.portal_form.keys():
            lab_building_days = lab_building_days + 0.5
            lab_test_days = lab_test_days + 0.5
        if "iseintegration" in self.portal_form.keys():
            lab_building_days = lab_building_days + 0.5
            lab_test_days = lab_test_days + 0.5

        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            # fill the workshop days
            self.ws["D42"] = lab_building_days * leader_effort
            self.ws["E42"] = lab_building_days * (1 - leader_effort)

            # fill the document creation days
            self.ws["D43"] = lab_test_days * leader_effort
            self.ws["E43"] = lab_test_days * (1 - leader_effort)
        else:
            self.ws["E42"] = lab_building_days

            self.ws["E43"] = lab_test_days

    def stw_implementation_testing_phase_editor(self):
        '''
                        according to the portal form value to fill in the workdays in LOE spreadsheet
                        :return:
        '''
        installation_days = 0.5
        basic_configuration_days = 2

        leader_effort = 0
        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100

        if "endpoint" in self.portal_form.keys():
            installation_days = installation_days + 0
            basic_configuration_days = basic_configuration_days + 1
        if "slic" in self.portal_form.keys():
            installation_days = installation_days + 0
            basic_configuration_days = basic_configuration_days + 0.5
        if "sal" in self.portal_form.keys():
            installation_days = installation_days + 0
            basic_configuration_days = basic_configuration_days + 0.5
        if "3rd" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D53"] = 2 * leader_effort
                self.ws["E53"] = 2 * (1 - leader_effort)
            else:
                self.ws["E53"] = 2
        if "api" in self.portal_form.keys():
            pass
        if "udpredirector" in self.portal_form.keys():
            installation_days = installation_days + 0.5
            basic_configuration_days = basic_configuration_days + 0.5
        if "fs" in self.portal_form.keys():
            installation_days = installation_days + 0.5
            basic_configuration_days = basic_configuration_days + 0.5
        if "iseintegration" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D51"] = 1 * leader_effort
                self.ws["E51"] = 1 * (1 - leader_effort)
            else:
                self.ws["E51"] = 1

        if self.portal_form["leader"] != str(0):
            self.ws["D49"] = installation_days * leader_effort
            self.ws["E49"] = installation_days * (1 - leader_effort)
            self.ws["D50"] = basic_configuration_days * leader_effort
            self.ws["E50"] = basic_configuration_days * (1 - leader_effort)
        else:
            self.ws["E49"] = installation_days

            self.ws["E50"] = basic_configuration_days

    def stw_kt_testing_phase_editor(self):
        '''
                        according to the portal form value to fill in the workdays in LOE spreadsheet
                        :return:
        '''
        deck_preparation_days = 4
        kt_days = 2

        if "endpoint" in self.portal_form.keys():
            deck_preparation_days = deck_preparation_days + 0.5
            kt_days = kt_days + 0.5
        if "slic" in self.portal_form.keys():
            deck_preparation_days = deck_preparation_days + 0.5
            kt_days = kt_days + 0.5
        if "sal" in self.portal_form.keys():
            deck_preparation_days = deck_preparation_days + 0.5
            kt_days = kt_days + 0.5
        if "3rd" in self.portal_form.keys():
            deck_preparation_days = deck_preparation_days + 1
            kt_days = kt_days + 1
        if "api" in self.portal_form.keys():
            deck_preparation_days = deck_preparation_days + 1
            kt_days = kt_days + 1
        if "iseintegration" in self.portal_form.keys():
            deck_preparation_days = deck_preparation_days + 1
            kt_days = kt_days + 1

        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D58"] = deck_preparation_days * leader_effort
            self.ws["E58"] = deck_preparation_days * (1 - leader_effort)

            self.ws["D59"] = kt_days * leader_effort
            self.ws["E59"] = kt_days * (1 - leader_effort)
        else:
            self.ws["E58"] = deck_preparation_days

            self.ws["E59"] = kt_days

    def stw_tunning_phase_editor(self):
        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            if self.portal_form["tunning"] == "1m":
                self.ws["D64"] = 15 * leader_effort
                self.ws["E64"] = 15 * (1 - leader_effort)
            elif self.portal_form["tunning"] == "3m":
                self.ws["D64"] = 40 * leader_effort
                self.ws["E64"] = 40 * (1 - leader_effort)
            elif self.portal_form["tunning"] == "6m":
                self.ws["D64"] = 80 * leader_effort
                self.ws["E64"] = 80 * (1 - leader_effort)
            else:
                self.ws["D64"] = 5 * leader_effort
                self.ws["E64"] = 5 * (1 - leader_effort)
        else:
            if self.portal_form["tunning"] == "1m":
                self.ws["E64"] = 15
            elif self.portal_form["tunning"] == "3m":
                self.ws["E64"] = 40
            elif self.portal_form["tunning"] == "6m":
                self.ws["E64"] = 80
            else:
                self.ws["E64"] = 5

# Firepower Editor

    def fp_requirement_phase_editor(self):
        '''
                according to the portal form value to fill in the workdays in LOE spreadsheet
                :return:
        '''
        workshop_days = 0
        document_creation_days = 0
        if self.portal_form["sdaornot"] == "yes":
            workshop_days = workshop_days + 1
        else:
            pass
        if self.portal_form["aciornot"] == "yes":
            workshop_days = workshop_days + 1
        else:
            pass

        if self.portal_form["deploymentmethod"] == "basicfw":
            workshop_days = workshop_days + 0.5
            document_creation_days = document_creation_days + 1
        elif self.portal_form["deploymentmethod"] == "threat":
            workshop_days = workshop_days + 1
            document_creation_days = document_creation_days + 1
        elif self.portal_form["deploymentmethod"] == "malware":
            workshop_days = workshop_days + 1
            document_creation_days = document_creation_days + 1
        elif self.portal_form["deploymentmethod"] == "ravpn":
            workshop_days = workshop_days + 1
            document_creation_days = document_creation_days + 1
        else:
            workshop_days = workshop_days + 2
            document_creation_days = document_creation_days + 2

        if "3rdparty" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
        if "trustsec" in self.portal_form.keys():
            workshop_days = workshop_days + 1
        if "aaa" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
        if "sslencryption" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
        if "onpremmalware" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
        if "auto" in self.portal_form.keys():
            workshop_days = workshop_days + 1
        if "autotest" in self.portal_form.keys():
            workshop_days = workshop_days + 2
        if "datamig" in self.portal_form.keys():
            workshop_days = workshop_days + 0

        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D17"] = workshop_days * leader_effort
            self.ws["E17"] = workshop_days * (1 - leader_effort)

            self.ws["D18"] = document_creation_days * leader_effort
            self.ws["E18"] = document_creation_days * (1 - leader_effort)
        else:
            self.ws["E17"] = workshop_days
            self.ws["E18"] = document_creation_days

    def fp_design_phase_editor(self):
        '''
                according to the portal form value to fill in the workdays in LOE spreadsheet
                :return:
        '''
        workshop_days = 1
        design_document_days = 0
        if self.portal_form["sdaornot"] == "yes":
            workshop_days = workshop_days + 1
            design_document_days = design_document_days + 1
        else:
            pass

        if self.portal_form["aciornot"] == "yes":
            workshop_days = workshop_days + 1
            design_document_days = design_document_days + 1
        else:
            pass

        if self.portal_form["deploymentmethod"] == "basicfw":
            design_document_days = design_document_days + 2
        elif self.portal_form["deploymentmethod"] == "threat":
            design_document_days = design_document_days + 2.5
        elif self.portal_form["deploymentmethod"] == "malware":
            design_document_days = design_document_days + 2.5
        elif self.portal_form["deploymentmethod"] == "ravpn":
            design_document_days = design_document_days + 3
        else:
            design_document_days = design_document_days + 4

        if "3rdparty" in self.portal_form.keys():
            design_document_days = design_document_days + 0.5
        if "trustsec" in self.portal_form.keys():
            design_document_days = design_document_days + 1
        if "aaa" in self.portal_form.keys():
            design_document_days = design_document_days + 1
        if "sslencryption" in self.portal_form.keys():
            design_document_days = design_document_days + 0.5
        if "onpremmalware" in self.portal_form.keys():
            design_document_days = design_document_days + 1

        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D21"] = workshop_days * leader_effort
            self.ws["E21"] = workshop_days * (1 - leader_effort)

            self.ws["D22"] = design_document_days * leader_effort
            self.ws["E22"] = design_document_days * (1 - leader_effort)
        else:
            self.ws["E21"] = workshop_days
            self.ws["E22"] = design_document_days

    def fp_nip_phase_editor(self):
        '''
            according to the portal form value to fill in the workdays in LOE spreadsheet
            :return:
        '''
        documentation_days = 0
        if self.portal_form["sdaornot"] == "yes":
            documentation_days = documentation_days + 1
        else:
            pass

        if self.portal_form["aciornot"] == "yes":
            documentation_days = documentation_days + 1
        else:
            pass

        if self.portal_form["deploymentmethod"] == "basicfw":
            documentation_days = documentation_days + 2
        elif self.portal_form["deploymentmethod"] == "threat":
            documentation_days = documentation_days + 3
        elif self.portal_form["deploymentmethod"] == "malware":
            documentation_days = documentation_days + 3
        elif self.portal_form["deploymentmethod"] == "ravpn":
            documentation_days = documentation_days + 3
        else:
            documentation_days = documentation_days + 4

        if "3rdparty" in self.portal_form.keys():
            documentation_days = documentation_days + 0.5
        if "trustsec" in self.portal_form.keys():
            documentation_days = documentation_days + 2
        if "aaa" in self.portal_form.keys():
            documentation_days = documentation_days + 0.5
        if "sslencryption" in self.portal_form.keys():
            documentation_days = documentation_days + 1
        if "onpremmalware" in self.portal_form.keys():
            documentation_days = documentation_days + 2

        # fill the document creation days
        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D29"] = documentation_days * leader_effort
            self.ws["E29"] = documentation_days * (1 - leader_effort)
        else:
            self.ws["E29"] = documentation_days

    def fp_nrfu_phase_editor(self):
        '''
                    according to the portal form value to fill in the workdays in LOE spreadsheet
                    :return:
        '''
        documentation_days = 0
        if self.portal_form["sdaornot"] == "yes":
            documentation_days = documentation_days + 1
        else:
            pass

        if self.portal_form["aciornot"] == "yes":
            documentation_days = documentation_days + 1
        else:
            pass

        if self.portal_form["deploymentmethod"] == "basicfw":
            documentation_days = documentation_days + 1
        elif self.portal_form["deploymentmethod"] == "threat":
            documentation_days = documentation_days + 2
        elif self.portal_form["deploymentmethod"] == "malware":
            documentation_days = documentation_days + 2
        elif self.portal_form["deploymentmethod"] == "ravpn":
            documentation_days = documentation_days + 2
        else:
            documentation_days = documentation_days + 3

        if "3rdparty" in self.portal_form.keys():
            documentation_days = documentation_days + 0.5
        if "trustsec" in self.portal_form.keys():
            documentation_days = documentation_days + 0.5
        if "aaa" in self.portal_form.keys():
            documentation_days = documentation_days + 0.5
        if "sslencryption" in self.portal_form.keys():
            documentation_days = documentation_days + 0.5
        if "onpremmalware" in self.portal_form.keys():
            documentation_days = documentation_days + 0.5

        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D35"] = documentation_days * leader_effort
            self.ws["E35"] = documentation_days * (1 - leader_effort)
        else:
            self.ws["E35"] = documentation_days

        if "autotest" in self.portal_form.keys():
            self.ws["D38"] = 10

    def fp_lab_testing_phase_editor(self):
        '''
                    according to the portal form value to fill in the workdays in LOE spreadsheet
                    :return:
        '''
        lab_building_days = 0
        test_execution_days = 0

        leader_effort = 0
        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100

        if self.portal_form["sdaornot"] == "yes":
            pass
        else:
            pass

        if self.portal_form["aciornot"] == "yes":
            pass
        else:
            pass

        if self.portal_form["deploymentmethod"] == "basicfw":
            lab_building_days = lab_building_days + 0
            test_execution_days = test_execution_days + 1
        elif self.portal_form["deploymentmethod"] == "threat":
            lab_building_days = lab_building_days + 1
            test_execution_days = test_execution_days + 1.5
        elif self.portal_form["deploymentmethod"] == "malware":
            lab_building_days = lab_building_days + 1
            test_execution_days = test_execution_days + 1.5
        elif self.portal_form["deploymentmethod"] == "ravpn":
            lab_building_days = lab_building_days + 2
            test_execution_days = test_execution_days + 2
        else:
            lab_building_days = lab_building_days + 2.5
            test_execution_days = test_execution_days + 2.5

        if "3rdparty" in self.portal_form.keys():
            lab_building_days = lab_building_days + 2
            if self.portal_form["leader"] != str(0):
                self.ws["D44"] = 2 * leader_effort
                self.ws["E44"] = 2 * (1 - leader_effort)
            else:
                self.ws["E44"] = 2
        if "trustsec" in self.portal_form.keys():
            lab_building_days = lab_building_days + 2
            test_execution_days = test_execution_days + 2
        if "aaa" in self.portal_form.keys():
            lab_building_days = lab_building_days + 1
            test_execution_days = test_execution_days + 1
        if "sslencryption" in self.portal_form.keys():
            lab_building_days = lab_building_days + 1
            test_execution_days = test_execution_days + 1
        if "onpremmalware" in self.portal_form.keys():
            lab_building_days = lab_building_days + 1
            test_execution_days = test_execution_days + 1
        if "auto" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D45"] = 10 * leader_effort
                self.ws["E45"] = 10 * (1 - leader_effort)
            else:
                self.ws["E45"] = 10
        if "autotest" in self.portal_form.keys():
            test_execution_days = test_execution_days + 5
        if "datamig" in self.portal_form.keys():
            lab_building_days = lab_building_days + 1

        if self.portal_form["leader"] != str(0):
            self.ws["D42"] = lab_building_days * leader_effort
            self.ws["E42"] = lab_building_days * (1 - leader_effort)
            self.ws["D43"] = test_execution_days * leader_effort
            self.ws["E43"] = test_execution_days * (1 - leader_effort)
        else:
            self.ws["E42"] = lab_building_days
            self.ws["E43"] = test_execution_days

    def fp_implementation_phase_editor(self):
        '''
                    according to the portal form value to fill in the workdays in LOE spreadsheet
                    :return:
        '''
        basic_configuration_days = 0
        leader_effort = 0
        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100

        if self.portal_form["sdaornot"] == "yes":
            basic_configuration_days = basic_configuration_days + 3
        else:
            pass

        if self.portal_form["aciornot"] == "yes":
            basic_configuration_days = basic_configuration_days + 3
        else:
            pass

        if self.portal_form["leader"] != str(0):
            self.ws["D50"] = basic_configuration_days * leader_effort
            self.ws["E50"] = basic_configuration_days * (1 - leader_effort)
        else:
            self.ws["E50"] = basic_configuration_days

        if self.portal_form["deploymentmethod"] == "basicfw":
            basic_configuration_days = basic_configuration_days + 3
            if self.portal_form["leader"] != str(0):
                self.ws["D50"] = basic_configuration_days * leader_effort
                self.ws["E50"] = basic_configuration_days * (1 - leader_effort)
            else:
                self.ws["E50"] = basic_configuration_days
        elif self.portal_form["deploymentmethod"] == "threat":
            if self.portal_form["leader"] != str(0):
                self.ws["D51"] = 2 * leader_effort
                self.ws["E51"] = 2 * (1 - leader_effort)
            else:
                self.ws["E51"] = 2
        elif self.portal_form["deploymentmethod"] == "malware":
            if self.portal_form["leader"] != str(0):
                self.ws["D52"] = 1 * leader_effort
                self.ws["E52"] = 1 * (1 - leader_effort)
            else:
                self.ws["E52"] = 1
        elif self.portal_form["deploymentmethod"] == "ravpn":
            if self.portal_form["leader"] != str(0):
                self.ws["D57"] = 3 * leader_effort
                self.ws["E57"] = 3 * (1 - leader_effort)
            else:
                self.ws["E57"] = 3
        else:
            if self.portal_form["leader"] != str(0):
                self.ws["D57"] = 5 * leader_effort
                self.ws["E57"] = 5 * (1 - leader_effort)
            else:
                self.ws["E57"] = 5

        if "3rdparty" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D59"] = 2 * leader_effort
                self.ws["E59"] = 2 * (1 - leader_effort)
            else:
                self.ws["E59"] = 2
        if "trustsec" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D56"] = 2 * leader_effort
                self.ws["E56"] = 2 * (1 - leader_effort)
            else:
                self.ws["E56"] = 2
        if "aaa" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D55"] = 1 * leader_effort
                self.ws["E55"] = 1 * (1 - leader_effort)
            else:
                self.ws["E55"] = 1
        if "sslencryption" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D54"] = 2 * leader_effort
                self.ws["E54"] = 2 * (1 - leader_effort)
            else:
                self.ws["E54"] = 2
        if "onpremmalware" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D53"] = 2 * leader_effort
                self.ws["E53"] = 2 * (1 - leader_effort)
            else:
                self.ws["E53"] = 2
        if "datamig" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D58"] = 1 * leader_effort
                self.ws["E58"] = 1 * (1 - leader_effort)
            else:
                self.ws["E58"] = 1

    def fp_kt_phase_editor(self):
        '''
            according to the portal form value to fill in the workdays in LOE spreadsheet
            :return:
        '''
        kt_document_days = 0
        training_days = 0
        if self.portal_form["sdaornot"] == "yes":
            pass
        else:
            pass

        if self.portal_form["aciornot"] == "yes":
            pass
        else:
            pass

        if self.portal_form["deploymentmethod"] == "basicfw":
            kt_document_days = kt_document_days + 2
            training_days = training_days + 0.5
        elif self.portal_form["deploymentmethod"] == "threat":
            kt_document_days = kt_document_days + 3
            training_days = training_days + 0.5
        elif self.portal_form["deploymentmethod"] == "malware":
            kt_document_days = kt_document_days + 4
            training_days = training_days + 1
        elif self.portal_form["deploymentmethod"] == "ravpn":
            kt_document_days = kt_document_days + 3
            training_days = training_days + 1.5
        else:
            kt_document_days = kt_document_days + 5
            training_days = training_days + 2

        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D64"] = kt_document_days * leader_effort
            self.ws["E64"] = kt_document_days * (1 - leader_effort)

            self.ws["D65"] = training_days * leader_effort
            self.ws["E65"] = training_days * (1 - leader_effort)
        else:
            self.ws["E64"] = kt_document_days
            self.ws["E65"] = training_days

# ISE Editor

    def ise_requirement_phase_editor(self):
        '''
        according to the portal form value to fill in the workdays in LOE spreadsheet
        :return:
        '''
        workshop_days = 0
        document_creation_days = 0
        if self.portal_form["sdaornot"] == "yes":
            workshop_days = workshop_days + 1
        else:
            pass

        if self.portal_form["deploymentmethod"] == "aaa":
            workshop_days = workshop_days + 1
            document_creation_days = document_creation_days + 1
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            workshop_days = workshop_days + 2
            document_creation_days = document_creation_days + 1
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            workshop_days = workshop_days + 2
            document_creation_days = document_creation_days + 1
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            workshop_days = workshop_days + 3
            document_creation_days = document_creation_days + 2
        else:
            workshop_days = workshop_days + 3
            document_creation_days = document_creation_days + 2

        if "3rdparty" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
        if "trustsec" in self.portal_form.keys():
            workshop_days = workshop_days + 1
        if "eapfast" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
        if "advancedguest" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5

        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D17"] = workshop_days * leader_effort
            self.ws["E17"] = workshop_days * (1 - leader_effort)
            self.ws["D18"] = document_creation_days * leader_effort
            self.ws["E18"] = document_creation_days * (1 - leader_effort)
        else:
            self.ws["E17"] = workshop_days
            self.ws["E18"] = document_creation_days

    def ise_design_phase_editor(self):
        '''
        according to the portal form value to fill in the workdays in LOE spreadsheet
        :return:
        '''
        workshop_days = 1
        design_document_days = 0
        if self.portal_form["sdaornot"] == "yes":
            workshop_days = workshop_days + 0

        if self.portal_form["deploymentmethod"] == "aaa":
            design_document_days = design_document_days + 4
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            design_document_days = design_document_days + 4.5
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            design_document_days = design_document_days + 5
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            workshop_days = workshop_days + 1
            design_document_days = design_document_days + 5.5
        else:
            workshop_days = workshop_days + 1
            design_document_days = design_document_days + 6

        if "3rdparty" in self.portal_form.keys():
            design_document_days = design_document_days + 0.5
        if "trustsec" in self.portal_form.keys():
            design_document_days = design_document_days + 1
        if "eapfast" in self.portal_form.keys():
            design_document_days = design_document_days + 1
        if "advancedguest" in self.portal_form.keys():
            design_document_days = design_document_days + 0.5

        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D21"] = workshop_days * leader_effort
            self.ws["E21"] = workshop_days * (1 - leader_effort)
            self.ws["D22"] = design_document_days * leader_effort
            self.ws["E22"] = design_document_days * (1 - leader_effort)
        else:
            self.ws["E21"] = workshop_days
            self.ws["E22"] = design_document_days

    def ise_nip_phase_editor(self):
        '''
            according to the portal form value to fill in the workdays in LOE spreadsheet
            :return:
        '''
        documentation_days = 0
        if self.portal_form["sdaornot"] == "yes":
            documentation_days = documentation_days + 0

        if self.portal_form["deploymentmethod"] == "aaa":
            documentation_days = documentation_days + 4
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            documentation_days = documentation_days + 5
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            documentation_days = documentation_days + 6
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            documentation_days = documentation_days + 7
        else:
            documentation_days = documentation_days + 8

        if "3rdparty" in self.portal_form.keys():
            documentation_days = documentation_days + 0.5
        if "trustsec" in self.portal_form.keys():
            documentation_days = documentation_days + 1
        if "eapfast" in self.portal_form.keys():
            documentation_days = documentation_days + 1
        if "advancedguest" in self.portal_form.keys():
            documentation_days = documentation_days + 0.5

        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D29"] = documentation_days * leader_effort
            self.ws["E29"] = documentation_days * (1 - leader_effort)
        else:
            self.ws["E29"] = documentation_days

    def ise_nruf_phase_editor(self):
        '''
            according to the portal form value to fill in the workdays in LOE spreadsheet
            :return:
        '''
        documentation_days = 0
        if self.portal_form["sdaornot"] == "yes":
            documentation_days = documentation_days + 0

        if self.portal_form["deploymentmethod"] == "aaa":
            documentation_days = documentation_days + 4
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            documentation_days = documentation_days + 5
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            documentation_days = documentation_days + 5
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            documentation_days = documentation_days + 5
        else:
            documentation_days = documentation_days + 5

        if "3rdparty" in self.portal_form.keys():
            documentation_days = documentation_days + 0
        if "trustsec" in self.portal_form.keys():
            documentation_days = documentation_days + 0
        if "eapfast" in self.portal_form.keys():
            documentation_days = documentation_days + 0
        if "advancedguest" in self.portal_form.keys():
            documentation_days = documentation_days + 0

        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D35"] = documentation_days * leader_effort
            self.ws["E35"] = documentation_days * (1 - leader_effort)
        else:
            self.ws["E35"] = documentation_days

        if "autotest" in self.portal_form.keys():
            self.ws["D38"] = 10

    def ise_lab_testing_phase_editor(self):
        '''
            according to the portal form value to fill in the workdays in LOE spreadsheet
            :return:
        '''
        lab_building_days = 0
        test_execution_days = 0

        leader_effort = 0
        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100

        if self.portal_form["sdaornot"] == "yes":
            pass

        if self.portal_form["deploymentmethod"] == "aaa":
            lab_building_days = lab_building_days + 0
            test_execution_days = test_execution_days + 1
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            lab_building_days = lab_building_days + 1
            test_execution_days = test_execution_days + 1.5
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            lab_building_days = lab_building_days + 1
            test_execution_days = test_execution_days + 1.5
            if self.portal_form["leader"] != str(0):
                self.ws["D44"] = 1 * leader_effort
                self.ws["E44"] = 1 * (1 - leader_effort)
            else:
                self.ws["E44"] = 1
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            lab_building_days = lab_building_days + 2
            test_execution_days = test_execution_days + 2
            test_execution_days = test_execution_days + 2.5
            if self.portal_form["leader"] != str(0):
                self.ws["D44"] = 1 * leader_effort
                self.ws["E44"] = 1 * (1 - leader_effort)
            else:
                self.ws["E44"] = 1
        else:
            lab_building_days = lab_building_days + 2.5
            test_execution_days = test_execution_days + 2.5
            if self.portal_form["leader"] != str(0):
                self.ws["D44"] = 1 * leader_effort
                self.ws["E44"] = 1 * (1 - leader_effort)
            else:
                self.ws["E44"] = 1

        if "3rdparty" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D45"] = 0.5 * leader_effort
                self.ws["E45"] = 0.5 * (1 - leader_effort)
            else:
                self.ws["E45"] = 0.5
        if "datamig" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D46"] = 0.5 * leader_effort
                self.ws["E46"] = 0.5 * (1 - leader_effort)
            else:
                self.ws["E46"] = 0.5
        if "auto" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D47"] = 2 * leader_effort
                self.ws["E47"] = 2 * (1 - leader_effort)
            else:
                self.ws["E47"] = 2

        # lab building days
        if self.portal_form["leader"] != str(0):
            self.ws["D42"] = lab_building_days * leader_effort
            self.ws["E42"] = lab_building_days * (1 - leader_effort)
        else:
            self.ws["E42"] = lab_building_days

        # testing days
        if self.portal_form["leader"] != str(0):
            self.ws["D43"] = test_execution_days * leader_effort
            self.ws["E43"] = test_execution_days * (1 - leader_effort)
        else:
            self.ws["E43"] = test_execution_days

    def ise_implementation_phase_editor(self):
        '''
            according to the portal form value to fill in the workdays in LOE spreadsheet
            :return:
        '''
        leader_effort = 0
        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100

        if self.portal_form["sdaornot"] == "yes":
            pass

        if self.portal_form["deploymentmethod"] == "aaa":
            if self.portal_form["leader"] != str(0):
                self.ws["D52"] = 3 * leader_effort
                self.ws["E52"] = 3 * (1 - leader_effort)
            else:
                self.ws["E52"] = 3
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            if self.portal_form["leader"] != str(0):
                self.ws["D52"] = 3 * leader_effort
                self.ws["E52"] = 3 * (1 - leader_effort)
                self.ws["D53"] = 0.5 * leader_effort
                self.ws["E53"] = 0.5 * (1 - leader_effort)
            else:
                self.ws["E52"] = 3
                self.ws["E53"] = 0.5
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            if self.portal_form["leader"] != str(0):
                self.ws["D52"] = 3 * leader_effort
                self.ws["E52"] = 3 * (1 - leader_effort)
                self.ws["D54"] = 0.5 * leader_effort
                self.ws["E54"] = 0.5 * (1 - leader_effort)
            else:
                self.ws["E52"] = 3
                self.ws["E54"] = 0.5
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            if self.portal_form["leader"] != str(0):
                self.ws["D52"] = 3 * leader_effort
                self.ws["E52"] = 3 * (1 - leader_effort)
                self.ws["D54"] = 0.5 * leader_effort
                self.ws["E54"] = 0.5 * (1 - leader_effort)
                self.ws["D55"] = 0.5 * leader_effort
                self.ws["E55"] = 0.5 * (1 - leader_effort)
            else:
                self.ws["E52"] = 3
                self.ws["E54"] = 0.5
                self.ws["E55"] = 0.5
        else:
            if self.portal_form["leader"] != str(0):
                self.ws["D52"] = 3 * leader_effort
                self.ws["E52"] = 3 * (1 - leader_effort)
                self.ws["D53"] = 0.5 * leader_effort
                self.ws["E53"] = 0.5 * (1 - leader_effort)
                self.ws["D54"] = 0.5 * leader_effort
                self.ws["E54"] = 0.5 * (1 - leader_effort)
                self.ws["D55"] = 0.5 * leader_effort
                self.ws["E55"] = 0.5 * (1 - leader_effort)
            else:
                self.ws["E52"] = 3
                self.ws["E53"] = 0.5
                self.ws["E54"] = 0.5
                self.ws["E55"] = 0.5

        if "3rdparty" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D57"] = 0.5 * leader_effort
                self.ws["E57"] = 0.5 * (1 - leader_effort)
            else:
                self.ws["E57"] = 0.5

        if "advanceguest" in self.portal_form.keys():
            if self.portal_form["leader"] != str(0):
                self.ws["D56"] = 0.5 * leader_effort
                self.ws["E56"] = 0.5 * (1 - leader_effort)
            else:
                self.ws["E56"] = 0.5

    def ise_kt_phase(self):
        '''
            according to the portal form value to fill in the workdays in LOE spreadsheet
            :return:
        '''
        kt_document_days = 0
        training_days = 0
        if self.portal_form["sdaornot"] == "yes":
            pass

        if self.portal_form["deploymentmethod"] == "aaa":
            kt_document_days = 3
            training_days = 0.5
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            kt_document_days = 4
            training_days = 1
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            kt_document_days = 4
            training_days = 1
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            kt_document_days = 4
            training_days = 1
        else:
            kt_document_days = 5
            training_days = 2

        # fill in the kt document days
        if self.portal_form["leader"] != str(0):
            leader_effort = int(self.portal_form["leader"]) / 100
            self.ws["D62"] = kt_document_days * leader_effort
            self.ws["E62"] = kt_document_days * (1 - leader_effort)
            self.ws["D63"] = training_days * leader_effort
            self.ws["E63"] = training_days * (1 - leader_effort)
        else:
            self.ws["E63"] = training_days

# Ad Hoc Feature
    def buffer_edit(self):
        '''
        edit the buffer cell
        '''
        try:
            buffer_value = int(self.portal_form["buffer"])
            cell_D = "D" + self.buffer_cell_line
            cell_E = "E" + self.buffer_cell_line
            if self.portal_form["leader"] != str(0):
                leader_effort = int(self.portal_form["leader"]) / 100
                self.ws[cell_D] = buffer_value * leader_effort
                self.ws[cell_E] = buffer_value * (1 - leader_effort)
                print("complete")
            else:
                self.ws[cell_E] = buffer_value
        except Exception as e:
            print("buffer empty")

    def empty_value(self):
        if self.portal_form["leader"] == str(0):
            for i in range(self.value_range[0], self.value_range[1]):
                self.ws.cell(row=i, column=4, value=0)
        else:
            print("no need empty the value")


# Save the spreadsheet
    def save_close_sheet(self, output_path):
        output_file_name = f"{self.portal_form['customer_name']}_Security_LoE.xlsx"
        self.wb.save(output_path + "/" + f"{self.portal_form['customer_name']}_Security_LoE.xlsx")
        return output_file_name
